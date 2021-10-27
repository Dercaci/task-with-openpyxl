import openpyxl


wb = openpyxl.load_workbook("des.xlsm")
ws = wb['Sheet1']

def change_cell(*args, **kwargs):
    row_count=1
    for row in ws.iter_rows(min_row=2, max_col=1):
        for cell in row:
        
            row_count+=1
            cell_list = cell.value
            cell_masive = cell_list.split()
            len_count = len(cell_masive)
            while len_count>1:
                len_count-=1
                n=0
                if cell_masive[n].isalpha()==False:
                    el = cell_masive[n]
                    cell_masive.remove(cell_masive[n])
                    cell_masive.append(el)
                    n+=1
            str1 = ""
            for ele in cell_masive:
                str1 += ele+' '

            cell1 = str1
            print(cell1)
            ws.cell(row=row_count, column=1).value = cell1
    return ws.cell

change_cell()
wb.save("des.xlsx")
